#!/usr/bin/env python3
"""Build an RCSA instrument workbook from a dataset module.

Usage:
    python3 build_xlsx.py <slug>        # populated worked example -> ../examples/<...>/instrument.xlsx
    python3 build_xlsx.py --blank OUT   # blank reusable template   -> OUT

A dataset module lives in ./data/<slug>_xlsx.py and exposes module-level
variables: META (dict), HOW_TO_USE (list[str]), COVER_META (list[(k,v)]),
SIGNOFF (list[(role,who,resp,date)]), RISKS, CONTROLS, ACTIONS, KRIS.
All rendering logic (styles, formulas, dropdowns, conditional formatting)
is identical for the template and every example — only the data differs.
"""
import sys, importlib, os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.worksheet.properties import PageSetupProperties

# ---------- palette ----------
NAVY = "1F3864"; NAVY2 = "2E4B7A"; SUB = "D9E1F2"; ALT = "F4F6FB"
CRIT_F, CRIT_T = "F4CCCC", "9C0006"
HIGH_F, HIGH_T = "FCE4D6", "9C4415"
MOD_F,  MOD_T  = "FFF2CC", "7F6000"
LOW_F,  LOW_T  = "E2EFDA", "375623"
INPUT_F = "FFFDE7"
GREY = "808080"
F = "Arial"

BLANK_ROWS = 18   # empty formula rows laid down in the reusable template

def font(sz=10, b=False, color="1A1A1A", italic=False):
    return Font(name=F, size=sz, bold=b, color=color, italic=italic)
def fill(c): return PatternFill("solid", fgColor=c)
thin = Side(style="thin", color="BFBFBF")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
def align(h="left", v="top", wrap=True): return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def band_cf(ws, rng):
    ws.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=['"Critical"'], fill=fill(CRIT_F), font=font(10,True,CRIT_T)))
    ws.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=['"High"'],     fill=fill(HIGH_F), font=font(10,True,HIGH_T)))
    ws.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=['"Moderate"'], fill=fill(MOD_F),  font=font(10,True,MOD_T)))
    ws.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=['"Low"'],      fill=fill(LOW_F),  font=font(10,True,LOW_T)))


def build(D, out_path, blank=False):
    META = D["META"]
    wb = Workbook()

    # =====================================================================
    # TAB 1 — COVER & SIGN-OFF
    # =====================================================================
    ws = wb.active; ws.title = "1. Cover & Sign-off"
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 34
    ws.column_dimensions["C"].width = 60
    ws.column_dimensions["D"].width = 28
    r = 2
    ws.cell(r,2,"RISK & CONTROL SELF-ASSESSMENT (RCSA)").font = font(18, True, NAVY); r+=1
    ws.cell(r,2,META["xlsx_masthead_2"]).font = font(11, True, NAVY2); r+=1
    ws.cell(r,2,META["xlsx_masthead_3"]).font = font(10, False, GREY, italic=True); r+=1
    ws.cell(r,2,META["xlsx_byline"]).font = font(9, False, GREY); r+=2

    def note(row, text, sz=9, color="404040", it=True, span=("B","D")):
        c = ws.cell(row,2,text); c.font = font(sz, False, color, italic=it); c.alignment = align(wrap=True)
        ws.merge_cells(f"{span[0]}{row}:{span[1]}{row}")
        ws.row_dimensions[row].height = 42
    note(r, META["cover_note"]); r+=2

    ws.cell(r,2,"How to use this instrument").font = font(11, True, NAVY); r+=1
    for line in D["HOW_TO_USE"]:
        c = ws.cell(r,2,line); c.font = font(9); c.alignment = align(wrap=True)
        ws.merge_cells(f"B{r}:D{r}"); ws.row_dimensions[r].height = 26; r+=1
    r+=1

    ws.cell(r,2,"Assessment record").font = font(11, True, NAVY); r+=1
    for k,v in D["COVER_META"]:
        a = ws.cell(r,2,k); a.fill=fill(SUB); a.font=font(9,True,NAVY); a.border=border; a.alignment=align("left","center")
        b = ws.cell(r,3,v); b.font=font(9); b.border=border; b.alignment=align("left","center",False)
        ws.cell(r,4).border=border; ws.merge_cells(f"C{r}:D{r}")
        r+=1
    r+=1

    ws.cell(r,2,"Governance & sign-off (three lines of defence)").font = font(11, True, NAVY); r+=1
    hdr = ["Line / role","Accountable role (enter name)","Responsibility","Sign-off date"]
    for i,h in enumerate(hdr):
        cell = ws.cell(r, 2+i, h); cell.fill=fill(NAVY); cell.font=font(9,True,"FFFFFF"); cell.border=border; cell.alignment=align("left","center")
    r+=1
    for role,who,resp,date in D["SIGNOFF"]:
        ws.cell(r,2,role).font=font(9,True); ws.cell(r,2).border=border; ws.cell(r,2).alignment=align("left","center",False)
        cwho=ws.cell(r,3,who); cwho.font=font(9); cwho.border=border; cwho.alignment=align("left","center",True); cwho.fill=fill(INPUT_F)
        ws.cell(r,4,resp).font=font(9); ws.cell(r,4).border=border; ws.cell(r,4).alignment=align("left","center",True)
        cdate=ws.cell(r,5,date); cdate.border=border; cdate.fill=fill(INPUT_F)
        r+=1
    ws.column_dimensions["E"].width = 16
    r+=1
    note(r, META["cover_legend"], span=("B","E"))
    ws.merge_cells(f"B{r}:E{r}")

    # =====================================================================
    # TAB 2 — METHODOLOGY & APPETITE  (identical for template & examples)
    # =====================================================================
    ws = wb.create_sheet("2. Methodology & Appetite")
    ws.sheet_view.showGridLines = False
    for k,v in {"A":3,"B":16,"C":20,"D":74}.items(): ws.column_dimensions[k].width = v
    r = 2
    ws.cell(r,2,"Methodology & Risk Appetite").font = font(16, True, NAVY); r+=2
    def tbl_title(row, t):
        ws.cell(row,2,t).font = font(11, True, NAVY); return row+1

    r = tbl_title(r, "A. Likelihood (L) — over a 12-month horizon")
    for i,h in enumerate(["Score","Rating","Definition"]):
        c=ws.cell(r,2+i,h); c.fill=fill(NAVY); c.font=font(9,True,"FFFFFF"); c.border=border; c.alignment=align("center" if i==0 else "left","center")
    r+=1
    for s,rt,d in [
     (1,"Rare","<5% in 12 months; no known precedent."),
     (2,"Unlikely","5–25%; could occur but not expected."),
     (3,"Possible","25–50%; may occur occasionally."),
     (4,"Likely","50–85%; expected in most circumstances."),
     (5,"Almost Certain",">85%; expected frequently, or already observed / recurring."),
    ]:
        ws.cell(r,2,s).alignment=align("center","center",False); ws.cell(r,2,s).font=font(9,True); ws.cell(r,2).border=border
        ws.cell(r,3,rt).font=font(9); ws.cell(r,3).border=border; ws.cell(r,3).alignment=align("left","center")
        ws.cell(r,4,d).font=font(9); ws.cell(r,4).border=border; ws.cell(r,4).alignment=align("left","center",True)
        r+=1
    r+=1

    r = tbl_title(r, "B. Impact (I) — highest applicable dimension governs")
    for i,h in enumerate(["Score","Rating","Illustrative anchors (safety/alignment · security · third-party & regulatory · reputational · operational)"]):
        c=ws.cell(r,2+i,h); c.fill=fill(NAVY); c.font=font(9,True,"FFFFFF"); c.border=border; c.alignment=align("center" if i==0 else "left","center",True)
    r+=1
    for s,rt,d in [
     (1,"Insignificant","Negligible; contained within a single workload; no external effect."),
     (2,"Minor","Limited, readily remediated; no external impact; immaterial cost."),
     (3,"Moderate","Notable internal impact; limited external or data exposure; local regulatory interest."),
     (4,"Major","Serious safety/security failure; third-party harm or data exposure; significant reputational/regulatory attention."),
     (5,"Severe","Loss of control of a capable system; material third-party breach; potential self-propagation or weight exfiltration; systemic reputational/regulatory harm."),
    ]:
        ws.cell(r,2,s).alignment=align("center","center",False); ws.cell(r,2,s).font=font(9,True); ws.cell(r,2).border=border
        ws.cell(r,3,rt).font=font(9); ws.cell(r,3).border=border; ws.cell(r,3).alignment=align("left","center")
        ws.cell(r,4,d).font=font(9); ws.cell(r,4).border=border; ws.cell(r,4).alignment=align("left","center",True)
        ws.row_dimensions[r].height=28
        r+=1
    r+=1

    r = tbl_title(r, "C. Inherent / Residual risk = L × I")
    for i,h in enumerate(["Score range","Band","Response posture"]):
        c=ws.cell(r,2+i,h); c.fill=fill(NAVY); c.font=font(9,True,"FFFFFF"); c.border=border; c.alignment=align("left","center")
    r+=1
    for rng,bandv,post,fc,tc in [
     ("1–4","Low","Accept / monitor.",LOW_F,LOW_T),
     ("5–9","Moderate","Manage via routine controls; periodic review.",MOD_F,MOD_T),
     ("10–15","High","Priority remediation; senior-management oversight.",HIGH_F,HIGH_T),
     ("16–25","Critical","Immediate action; executive / board attention; consider pausing activity.",CRIT_F,CRIT_T),
    ]:
        ws.cell(r,2,rng).font=font(9); ws.cell(r,2).border=border; ws.cell(r,2).alignment=align("center","center",False)
        b=ws.cell(r,3,bandv); b.fill=fill(fc); b.font=font(9,True,tc); b.border=border; b.alignment=align("center","center",False)
        ws.cell(r,4,post).font=font(9); ws.cell(r,4).border=border; ws.cell(r,4).alignment=align("left","center",True)
        r+=1
    r+=1

    r = tbl_title(r, "D. Control effectiveness (assessed separately for Design and Operating)")
    for i,h in enumerate(["Rating","Definition"]):
        c=ws.cell(r,2+i,h); c.fill=fill(NAVY); c.font=font(9,True,"FFFFFF"); c.border=border; c.alignment=align("left","center")
    ws.merge_cells(f"C{r}:D{r}")
    r+=1
    for rt,d in [
     ("Effective","Well designed / operating consistently to mitigate the risk."),
     ("Partially Effective","Addresses the risk in part; material weaknesses remain."),
     ("Ineffective","Present but does not mitigate the risk as intended."),
     ("Not Designed / Not Operating","No control designed, or the intended control is not in operation."),
    ]:
        ws.cell(r,2,rt).font=font(9,True); ws.cell(r,2).border=border; ws.cell(r,2).alignment=align("left","center",True)
        ws.cell(r,3,d).font=font(9); ws.cell(r,3).border=border; ws.cell(r,3).alignment=align("left","center",True); ws.merge_cells(f"C{r}:D{r}")
        r+=1
    r+=1

    r = tbl_title(r, "E. Risk appetite & tolerance")
    for line in [
     "Overall posture: AVERSE to loss of model control. The organisation has NO appetite for residual risk in the High or Critical band on any Model-Risk/Misalignment or containment-critical risk.",
     "Tolerance ceilings (maximum acceptable residual):",
     "   •  Model Risk / Misalignment and containment-critical Cyber (sandbox, persistence): residual must be LOW (score ≤ 4).",
     "   •  Other Security, Third-Party, Process, Monitoring and Governance risks: residual must be MODERATE or lower (score ≤ 9).",
     "Any residual above the applicable ceiling is a BREACH of appetite and requires a tracked action (see G). Residual is assessed on controls as they actually operate.",
    ]:
        c=ws.cell(r,2,line); c.font=font(9, ("Tolerance ceilings" in line or "Overall posture" in line)); c.alignment=align(wrap=True)
        ws.merge_cells(f"B{r}:D{r}"); ws.row_dimensions[r].height = 26 if len(line)>90 else 15; r+=1
    r+=1

    r = tbl_title(r, "F. Appetite status & G. Action priority (auto-calculated in the register)")
    for line in [
     "Appetite status:  Within appetite (residual ≤ ceiling)  ·  At tolerance (residual = ceiling)  ·  BREACH (residual > ceiling).",
     "Action priority for breaches:  P1 – Immediate (residual score ≥ 20)  ·  P2 – High (12–19)  ·  P3 – Medium (5–11).  Within-appetite risks are set to Monitor.",
    ]:
        c=ws.cell(r,2,line); c.font=font(9); c.alignment=align(wrap=True); ws.merge_cells(f"B{r}:D{r}"); ws.row_dimensions[r].height=26; r+=1

    # =====================================================================
    # TAB 3 — RISK REGISTER
    # =====================================================================
    ws = wb.create_sheet("3. Risk Register")
    ws.sheet_view.showGridLines = False
    cols = ["Risk ID","Sub-process / activity","Risk category (taxonomy)","Risk event (cause → event → consequence)",
            "Inh. L","Inh. I","Inh. score","Inh. band","Key controls","Ctrl design","Ctrl operating","Overall ctrl",
            "Res. L","Res. I","Res. score","Res. band","Appetite (max band)","Ceiling","Residual vs appetite","Action priority",
            "Risk owner (role)","KRI ref","Action ref","Basis / source"]
    widths = [8,20,20,46,6,6,7,9,12,13,13,12,6,6,7,9,13,7,17,15,22,7,8,26]
    for i,w in enumerate(widths): ws.column_dimensions[get_column_letter(i+1)].width = w
    ws.merge_cells("A1:X1"); t=ws.cell(1,1,META["xlsx_register_title"]); t.font=font(11,True,NAVY)
    hdr_row = 2
    for i,h in enumerate(cols):
        c=ws.cell(hdr_row,i+1,h); c.fill=fill(NAVY); c.font=font(8,True,"FFFFFF"); c.border=border; c.alignment=align("center","center",True)
    ws.row_dimensions[hdr_row].height=34

    def write_reg_row(row, rk):
        # rk: tuple (rid,sub,cat,evt,il,ii,ctrls,des,ope,rl,ri,appb,ceil,owner,kri,act,basis) or None for blank
        if rk is None:
            rid=sub=cat=evt=ctrls=des=ope=appb=owner=kri=act=basis=""
            il=ii=rl=ri=ceil=""
        else:
            (rid,sub,cat,evt,il,ii,ctrls,des,ope,rl,ri,appb,ceil,owner,kri,act,basis) = rk
        ws.cell(row,1,rid).font=font(9,True,NAVY)
        ws.cell(row,2,sub).font=font(9)
        ws.cell(row,3,cat).font=font(9)
        ws.cell(row,4,evt).font=font(8)
        ws.cell(row,5,il).font=font(9); ws.cell(row,5).fill=fill(INPUT_F)
        ws.cell(row,6,ii).font=font(9); ws.cell(row,6).fill=fill(INPUT_F)
        ws.cell(row,7,f"=IF(COUNT(E{row}:F{row})=2,E{row}*F{row},\"\")").font=font(9,True)
        ws.cell(row,8,f'=IF(G{row}="","",IF(G{row}>=16,"Critical",IF(G{row}>=10,"High",IF(G{row}>=5,"Moderate","Low"))))').font=font(9,True)
        ws.cell(row,9,ctrls).font=font(9)
        ws.cell(row,10,des).font=font(9); ws.cell(row,10).fill=fill(INPUT_F)
        ws.cell(row,11,ope).font=font(9); ws.cell(row,11).fill=fill(INPUT_F)
        ws.cell(row,12,f'=IF(AND(J{row}="",K{row}=""),"",IF(OR(J{row}="Not Designed",J{row}="Not Implemented",J{row}="Not Operating",K{row}="Not Operating"),"Not established",IF(AND(J{row}="Effective",K{row}="Effective"),"Effective",IF(OR(K{row}="Ineffective"),"Ineffective","Partially effective"))))').font=font(9)
        ws.cell(row,13,rl).font=font(9); ws.cell(row,13).fill=fill(INPUT_F)
        ws.cell(row,14,ri).font=font(9); ws.cell(row,14).fill=fill(INPUT_F)
        ws.cell(row,15,f"=IF(COUNT(M{row}:N{row})=2,M{row}*N{row},\"\")").font=font(9,True)
        ws.cell(row,16,f'=IF(O{row}="","",IF(O{row}>=16,"Critical",IF(O{row}>=10,"High",IF(O{row}>=5,"Moderate","Low"))))').font=font(9,True)
        ws.cell(row,17,appb).font=font(9); ws.cell(row,17).fill=fill(INPUT_F); ws.cell(row,17).alignment=align("center","center",False)
        ws.cell(row,18,ceil).font=font(9); ws.cell(row,18).fill=fill(INPUT_F); ws.cell(row,18).alignment=align("center","center",False)
        ws.cell(row,19,f'=IF(OR(O{row}="",R{row}=""),"",IF(O{row}>R{row},"BREACH",IF(O{row}=R{row},"At tolerance","Within appetite")))').font=font(9,True)
        ws.cell(row,20,f'=IF(S{row}="","",IF(S{row}="Within appetite","Monitor",IF(O{row}>=20,"P1 - Immediate",IF(O{row}>=12,"P2 - High","P3 - Medium"))))').font=font(9,True)
        ws.cell(row,21,owner).font=font(9)
        ws.cell(row,22,kri).font=font(9); ws.cell(row,22).alignment=align("center","center",False)
        ws.cell(row,23,act).font=font(9); ws.cell(row,23).alignment=align("center","center",False)
        ws.cell(row,24,basis).font=font(8, italic=True, color=GREY)
        for c in range(1,25):
            cell=ws.cell(row,c); cell.border=border
            if cell.alignment is None or cell.alignment.horizontal is None:
                cell.alignment=align("left","center",True)
            else:
                cell.alignment=Alignment(horizontal=cell.alignment.horizontal, vertical="center", wrap_text=True)
        ws.row_dimensions[row].height=54

    rows_data = D["RISKS"] if not blank else [None]*BLANK_ROWS
    row = hdr_row+1; first = row
    for rk in rows_data:
        write_reg_row(row, rk); row+=1
    last = row-1
    ws.freeze_panes = "B3"
    for col in [5,6,7,13,14,15,18]:
        for rr in range(first,last+1):
            ws.cell(rr,col).alignment=align("center","center",False)
    band_cf(ws, f"H{first}:H{last}")
    band_cf(ws, f"P{first}:P{last}")
    ws.conditional_formatting.add(f"S{first}:S{last}", CellIsRule(operator="equal", formula=['"BREACH"'], fill=fill(CRIT_F), font=font(9,True,CRIT_T)))
    ws.conditional_formatting.add(f"S{first}:S{last}", CellIsRule(operator="equal", formula=['"At tolerance"'], fill=fill(MOD_F), font=font(9,True,MOD_T)))
    ws.conditional_formatting.add(f"S{first}:S{last}", CellIsRule(operator="equal", formula=['"Within appetite"'], fill=fill(LOW_F), font=font(9,True,LOW_T)))
    ws.conditional_formatting.add(f"T{first}:T{last}", FormulaRule(formula=[f'ISNUMBER(SEARCH("P1",T{first}))'], fill=fill(CRIT_F), font=font(9,True,CRIT_T)))
    ws.conditional_formatting.add(f"T{first}:T{last}", FormulaRule(formula=[f'ISNUMBER(SEARCH("P2",T{first}))'], fill=fill(HIGH_F), font=font(9,True,HIGH_T)))
    ws.conditional_formatting.add(f"T{first}:T{last}", FormulaRule(formula=[f'ISNUMBER(SEARCH("P3",T{first}))'], fill=fill(MOD_F), font=font(9,True,MOD_T)))
    dv_li = DataValidation(type="list", formula1='"1,2,3,4,5"', allow_blank=True)
    dv_des = DataValidation(type="list", formula1='"Effective,Partially Effective,Ineffective,Not Designed"', allow_blank=True)
    dv_ope = DataValidation(type="list", formula1='"Effective,Partially Effective,Ineffective,Not Operating"', allow_blank=True)
    dv_app = DataValidation(type="list", formula1='"Low,Moderate,High,Critical"', allow_blank=True)
    for dv in (dv_li,dv_des,dv_ope,dv_app): ws.add_data_validation(dv)
    for rr in range(first,last+1):
        dv_li.add(ws.cell(rr,5)); dv_li.add(ws.cell(rr,6)); dv_li.add(ws.cell(rr,13)); dv_li.add(ws.cell(rr,14))
        dv_des.add(ws.cell(rr,10)); dv_ope.add(ws.cell(rr,11)); dv_app.add(ws.cell(rr,17))

    # =====================================================================
    # TAB 4 — CONTROL LIBRARY
    # =====================================================================
    ws = wb.create_sheet("4. Control Library")
    ws.sheet_view.showGridLines = False
    cols = ["Ctrl ID","Control objective / description","Related risks","Type","Nature","Key?","Frequency","Control owner (role)","Design","Operating","Overall","Deficiency note","Remediation direction"]
    widths=[8,40,12,12,13,6,15,24,14,14,14,42,42]
    for i,w in enumerate(widths): ws.column_dimensions[get_column_letter(i+1)].width=w
    ws.merge_cells("A1:M1"); ws.cell(1,1,META["xlsx_control_title"]).font=font(11,True,NAVY)
    for i,h in enumerate(cols):
        c=ws.cell(2,i+1,h); c.fill=fill(NAVY); c.font=font(8,True,"FFFFFF"); c.border=border; c.alignment=align("center","center",True)
    ws.row_dimensions[2].height=30

    def write_ctl_row(row, c):
        if c is None:
            vals=["","","","","","","","","","",None,"",""]
        else:
            (cid,obj,rel,typ,nat,key,freq,owner,des,ope,defnote,rem)=c
            vals=[cid,obj,rel,typ,nat,key,freq,owner,des,ope,None,defnote,rem]
        for i,v in enumerate(vals):
            cell=ws.cell(row,i+1)
            if i==10:
                cell.value=f'=IF(AND(I{row}="",J{row}=""),"",IF(OR(I{row}="Not Designed",I{row}="Not Operating",J{row}="Not Operating"),"Not established",IF(AND(I{row}="Effective",J{row}="Effective"),"Effective",IF(J{row}="Ineffective","Ineffective","Partially effective"))))'
            else:
                cell.value=v
            cell.border=border
            cell.font=font(9,True,NAVY) if i==0 else font(8) if i in (11,12) else font(9)
            cell.alignment=align("center" if i in (0,3,4,5,6) else "left","center",True)
            if i in (8,9): cell.fill=fill(INPUT_F)
        ws.row_dimensions[row].height=44

    cdata = D["CONTROLS"] if not blank else [None]*BLANK_ROWS
    row=3; first=row
    for c in cdata:
        write_ctl_row(row, c); row+=1
    last=row-1
    ws.freeze_panes="A3"
    for col in ("I","J","K"):
        rng=f"{col}{first}:{col}{last}"
        ws.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=['"Effective"'], fill=fill(LOW_F), font=font(9,True,LOW_T)))
        ws.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=['"Partially Effective"'], fill=fill(MOD_F), font=font(9,True,MOD_T)))
        ws.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=['"Ineffective"'], fill=fill(CRIT_F), font=font(9,True,CRIT_T)))
        ws.conditional_formatting.add(rng, FormulaRule(formula=[f'ISNUMBER(SEARCH("Not",{col}{first}))'], fill=fill(CRIT_F), font=font(9,True,CRIT_T)))
    dv_des2=DataValidation(type="list", formula1='"Effective,Partially Effective,Ineffective,Not Designed"', allow_blank=True)
    dv_ope2=DataValidation(type="list", formula1='"Effective,Partially Effective,Ineffective,Not Operating"', allow_blank=True)
    dv_key=DataValidation(type="list", formula1='"Y,N"', allow_blank=True)
    dv_typ=DataValidation(type="list", formula1='"Preventive,Detective,Corrective"', allow_blank=True)
    for dv in (dv_des2,dv_ope2,dv_key,dv_typ): ws.add_data_validation(dv)
    for rr in range(first,last+1):
        dv_des2.add(ws.cell(rr,9)); dv_ope2.add(ws.cell(rr,10)); dv_key.add(ws.cell(rr,6)); dv_typ.add(ws.cell(rr,4))

    # =====================================================================
    # TAB 5 — ACTION PLAN
    # =====================================================================
    ws = wb.create_sheet("5. Action Plan")
    ws.sheet_view.showGridLines=False
    cols=["Action ID","Addresses (risk / gap)","Finding (control gap)","Action","Owner (role)","Priority","Target window","Status","Target residual"]
    widths=[9,16,40,44,24,11,14,24,13]
    for i,w in enumerate(widths): ws.column_dimensions[get_column_letter(i+1)].width=w
    ws.merge_cells("A1:I1"); ws.cell(1,1,META["xlsx_action_title"]).font=font(11,True,NAVY)
    for i,h in enumerate(cols):
        c=ws.cell(2,i+1,h); c.fill=fill(NAVY); c.font=font(8,True,"FFFFFF"); c.border=border; c.alignment=align("center","center",True)
    adata = D["ACTIONS"] if not blank else [("",)*9]*BLANK_ROWS
    row=3; first=row
    for a in adata:
        for i,v in enumerate(a):
            cell=ws.cell(row,i+1,v); cell.border=border
            cell.font=font(9,True,NAVY) if i==0 else font(9)
            cell.alignment=align("center" if i in (0,5,6,8) else "left","center",True)
            if i in (7,8): cell.fill=fill(INPUT_F)
        ws.row_dimensions[row].height=40
        row+=1
    last=row-1
    ws.freeze_panes="A3"
    ws.conditional_formatting.add(f"F{first}:F{last}", FormulaRule(formula=[f'ISNUMBER(SEARCH("P1",F{first}))'], fill=fill(CRIT_F), font=font(9,True,CRIT_T)))
    ws.conditional_formatting.add(f"F{first}:F{last}", FormulaRule(formula=[f'ISNUMBER(SEARCH("P2",F{first}))'], fill=fill(HIGH_F), font=font(9,True,HIGH_T)))
    band_cf(ws, f"I{first}:I{last}")

    # =====================================================================
    # TAB 6 — KRI / KCI REGISTER
    # =====================================================================
    ws = wb.create_sheet("6. KRI-KCI Register")
    ws.sheet_view.showGridLines=False
    cols=["ID","Indicator","Type","Definition / measure","Threshold / appetite (target)","Current status (illustrative)","Owner (role)","Related risks"]
    widths=[7,34,8,40,30,26,24,14]
    for i,w in enumerate(widths): ws.column_dimensions[get_column_letter(i+1)].width=w
    ws.merge_cells("A1:H1"); ws.cell(1,1,"Key Risk & Control Indicators (KRI / KCI)  —  targets illustrative").font=font(11,True,NAVY)
    for i,h in enumerate(cols):
        c=ws.cell(2,i+1,h); c.fill=fill(NAVY); c.font=font(8,True,"FFFFFF"); c.border=border; c.alignment=align("center","center",True)
    kdata = D["KRIS"] if not blank else [("",)*8]*BLANK_ROWS
    row=3
    for k in kdata:
        for i,v in enumerate(k):
            cell=ws.cell(row,i+1,v); cell.border=border
            cell.font=font(9,True,NAVY) if i==0 else font(9)
            cell.alignment=align("center" if i in (0,2) else "left","center",True)
            if i==5: cell.fill=fill(INPUT_F)
        ws.row_dimensions[row].height=34
        row+=1
    ws.freeze_panes="A3"

    # =====================================================================
    # TAB 7 — PROFILE SUMMARY  (identical formulas for template & examples)
    # =====================================================================
    ws = wb.create_sheet("7. Profile Summary")
    ws.sheet_view.showGridLines=False
    for k,w in {"A":3,"B":26,"C":12,"D":12,"E":12,"F":12,"G":12,"H":4}.items(): ws.column_dimensions[k].width=w
    RR="'3. Risk Register'"
    r=2
    ws.cell(r,2,"Risk Profile Summary").font=font(16,True,NAVY); r+=2
    ws.cell(r,2,"Auto-calculated from the Risk Register.").font=font(9,italic=True,color=GREY); r+=2
    ws.cell(r,2,"Count by band").font=font(11,True,NAVY); r+=1
    ws.cell(r,2,"Band").font=font(9,True,"FFFFFF"); ws.cell(r,2).fill=fill(NAVY); ws.cell(r,2).border=border
    ws.cell(r,3,"Inherent").font=font(9,True,"FFFFFF"); ws.cell(r,3).fill=fill(NAVY); ws.cell(r,3).border=border; ws.cell(r,3).alignment=align("center","center",False)
    ws.cell(r,4,"Residual").font=font(9,True,"FFFFFF"); ws.cell(r,4).fill=fill(NAVY); ws.cell(r,4).border=border; ws.cell(r,4).alignment=align("center","center",False)
    r+=1
    band_start=r
    for bandv,fc,tc in [("Critical",CRIT_F,CRIT_T),("High",HIGH_F,HIGH_T),("Moderate",MOD_F,MOD_T),("Low",LOW_F,LOW_T)]:
        b=ws.cell(r,2,bandv); b.fill=fill(fc); b.font=font(9,True,tc); b.border=border
        ci=ws.cell(r,3,f'=COUNTIF({RR}!H:H,"{bandv}")'); ci.border=border; ci.alignment=align("center","center",False); ci.font=font(9)
        cr=ws.cell(r,4,f'=COUNTIF({RR}!P:P,"{bandv}")'); cr.border=border; cr.alignment=align("center","center",False); cr.font=font(9)
        r+=1
    tot=ws.cell(r,2,"Total"); tot.font=font(9,True); tot.border=border
    ws.cell(r,3,f"=SUM(C{band_start}:C{r-1})").border=border; ws.cell(r,3).alignment=align("center","center",False); ws.cell(r,3).font=font(9,True)
    ws.cell(r,4,f"=SUM(D{band_start}:D{r-1})").border=border; ws.cell(r,4).alignment=align("center","center",False); ws.cell(r,4).font=font(9,True)
    r+=2
    ws.cell(r,2,"Appetite & priority").font=font(11,True,NAVY); r+=1
    for label,formula,fc in [
     ("Risks breaching appetite",f'=COUNTIF({RR}!S:S,"BREACH")',CRIT_F),
     ("Within appetite",f'=COUNTIF({RR}!S:S,"Within appetite")',LOW_F),
     ("P1 – Immediate actions",f'=COUNTIF({RR}!T:T,"P1 - Immediate")',CRIT_F),
     ("P2 – High actions",f'=COUNTIF({RR}!T:T,"P2 - High")',HIGH_F),
     ("P3 – Medium actions",f'=COUNTIF({RR}!T:T,"P3 - Medium")',MOD_F),
    ]:
        a=ws.cell(r,2,label); a.font=font(9); a.border=border; a.fill=fill(fc)
        v=ws.cell(r,3,formula); v.border=border; v.alignment=align("center","center",False); v.font=font(9,True)
        ws.cell(r,4).border=border
        r+=1
    r+=2
    ws.cell(r,2,"Residual risk heat map (count of risks)").font=font(11,True,NAVY); r+=1
    ws.cell(r,2,"Impact ↓  /  Likelihood →").font=font(8,italic=True,color=GREY)
    for L in range(1,6):
        c=ws.cell(r,2+L,L); c.font=font(9,True,"FFFFFF"); c.fill=fill(NAVY); c.border=border; c.alignment=align("center","center",False)
    r+=1
    for I in range(5,0,-1):
        lab=ws.cell(r,2,I); lab.font=font(9,True,"FFFFFF"); lab.fill=fill(NAVY); lab.border=border; lab.alignment=align("center","center",False)
        for L in range(1,6):
            score=L*I
            fc = CRIT_F if score>=16 else HIGH_F if score>=10 else MOD_F if score>=5 else LOW_F
            cell=ws.cell(r,2+L,f'=COUNTIFS({RR}!$M:$M,{L},{RR}!$N:$N,{I})')
            cell.fill=fill(fc); cell.border=border; cell.alignment=align("center","center",False); cell.font=font(9,True)
        r+=1
    ws.cell(r+1,2,"Cells shaded by band (L×I). Populated from residual Likelihood/Impact in the register.").font=font(8,italic=True,color=GREY)

    title_rows = {"3. Risk Register":"1:2","4. Control Library":"1:2","5. Action Plan":"1:2","6. KRI-KCI Register":"1:2"}
    for ws in wb.worksheets:
        ws.page_setup.orientation = "landscape"
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
        ws.print_options.horizontalCentered = True
        ws.page_margins.left = ws.page_margins.right = 0.3
        ws.page_margins.top = ws.page_margins.bottom = 0.4
        if ws.title in title_rows:
            ws.print_title_rows = title_rows[ws.title]

    os.makedirs(os.path.dirname(out_path) or ".", exist_ok=True)
    wb.save(out_path)
    print("WROTE", out_path)


def load_dataset(slug):
    sys.path.insert(0, os.path.join(os.path.dirname(os.path.abspath(__file__)), "data"))
    mod = importlib.import_module(f"{slug}_xlsx")
    return mod.DATA

if __name__ == "__main__":
    here = os.path.dirname(os.path.abspath(__file__))
    if len(sys.argv) >= 3 and sys.argv[1] == "--blank":
        D = load_dataset(sys.argv[3]) if len(sys.argv) > 3 else load_dataset("template")
        build(D, sys.argv[2], blank=True)
    else:
        slug = sys.argv[1]
        D = load_dataset(slug)
        out = sys.argv[2] if len(sys.argv) > 2 else os.path.join(here, "..", "examples", D["META"]["dir"], "instrument.xlsx")
        build(D, out, blank=False)
